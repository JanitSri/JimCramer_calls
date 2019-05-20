'''
NAME: Janit Sriganeshaelankovan 
CREATED: Decmeber 17, 2018 - 18:56 (EDT)
GOAL: Analysis of Jim Cramer's Calls from Mad Money 
ENVIRONMENT: Base 
LAST UPDATE: May 19, 2019 - 08:53 (EDT)
'''

import os 
from bs4 import BeautifulSoup as soup
import requests 
import pandas as pd
import openpyxl
import re
import time 
import numpy as np
from yahooEODscraper import individual_security_eod
from yahoo_stats_scraper import get_stats, converter, more_stats 
import datetime
import json 


os.chdir()
os.getcwd()


# https://madmoney.thestreet.com/screener/index.cfm?showview=stocks&showrows=25000 -- most recent
# save the file as html and copy and paste to excel 

# path4 in the code is the folder(stocks_data) that holds the companies' stock data


'''FORMAT THE EXCEL SHEET [DATES COLUUMN]'''
wb = openpyxl.load_workbook('cleandata.xlsx')
sheets_names = wb.sheetnames

sheet = wb['Calls']
sheet.max_row
sheet.max_column

with open('date.txt', 'a') as f:
    for _row in sheet.iter_cols(min_row=2, min_col=2, max_row=2893, max_col=2):
        for cell in _row:
            value = str(cell.value).replace('2019', '2016').rstrip('00:00:00')
            print(value)
            f.write(value + '\n')

with open('date.txt', 'a') as f:
    for _row in sheet.iter_cols(min_row=2894, min_col=2, max_row=6925, max_col=2):
        for cell in _row:
            value = str(cell.value).replace('2019', '2017').rstrip('00:00:00')
            print(value)
            f.write(value + '\n')
        
with open('date.txt', 'a') as f:
    for _row in sheet.iter_cols(min_row=6926, min_col=2, max_row=10840, max_col=2):
        for cell in _row:
            value = str(cell.value).replace('2019', '2018').rstrip('00:00:00')
            print(value)
            f.write(value + '\n')

with open('date.txt', 'a') as f:
    for _row in sheet.iter_cols(min_row=10841, min_col=2, max_row=12049, max_col=2):
        for cell in _row:
            value = str(cell.value).rstrip('00:00:00')
            print(value)
            f.write(value + '\n')



'''READ IN THE DATA INTO PANDAS'''
df_cramer = pd.read_csv('cleandata.csv')
column_names = list(df_cramer.columns.values)
df_cramer.head()
df_cramer.tail()



'''PRE ANALYSIS'''
unique_companies = df_cramer['Ticker'].unique().tolist()
unique_companies = sorted(unique_companies)
len(unique_companies)
len(df_cramer)
df_unique_companies_count = df_cramer['Company'].value_counts()
max_company_calls = (df_unique_companies_count.idxmax(), df_unique_companies_count.max())
min_company_calls = (df_unique_companies_count.idxmin(), df_unique_companies_count.min())
df_unique_calls_count = df_cramer['Call'].value_counts()
df_unique_segements_count = df_cramer['Segement'].value_counts()



'''GET COMPANY EOD DATA'''
# Company EOD Data 
done = []
for idx, comp in enumerate(unique_companies):
    print('ON:', idx)
    individual_security_eod('01/01/1990', '22/04/2019', [comp])
    done.append(comp)

path, dirs, files = next(os.walk(path4))
file_count = len(files)



'''SECTOR INDUSTRY DATA'''
with open('ticker.txt', 'w') as f:
    f.write('Ticker')
    f.write('\n')
    for x in unique_companies:
        f.write(x)
        f.write('\n')

more_stats('ticker') 

# Map the Sector and Industry to orginal Data Frame 
df_sector_indus = pd.read_csv('ticker_moreinfo.csv')
df_sector_indus.columns

sector_dict = df_sector_indus.set_index('Ticker').to_dict()['Sector']
industry_dict = df_sector_indus.set_index('Ticker').to_dict()['Industry']

df_cramer['Sector'] = df_cramer['Ticker'].map(sector_dict)
df_cramer['Industry'] = df_cramer['Ticker'].map(industry_dict)

df_cramer["Sector"] = df_cramer["Sector"].replace({'",': 'NAN'})
df_cramer["Industry"] = df_cramer["Industry"].replace({'",': 'NAN'})

df_cramer.to_csv('map_test.csv')

df_unique_sector_count = df_cramer['Sector'].value_counts().drop(labels=['NAN'])
df_unique_industries_count = df_cramer['Industry'].value_counts().drop(labels=['NAN'])

df_unique_industries_count.plot(kind='barh')



'''CLEAN TICKER DATA (CONVERT EPOCH DATE TO M/D/YYYY)'''
df_cramer['Date'].head(10)
file_list = []

for root, dirs, files in os.walk(path4):  
    for filename in files:
        print(filename)
        file_list.append(filename)


for idx, file in enumerate(file_list[2:]):
    print(idx,file)
    
    ticker_data = []
    
    with open(path4+'\{}'.format(file), 'r') as f:
        for data in f.readlines():
            ticker_data.append(data.strip('\n').split(","))
            
    with open(path4+'\\'+file,'w'): pass  # earse the file

    with open(path4+'\{}'.format(file), 'w') as f:
        for td in ticker_data:
            td[0] = int(td[0])
            td[0] = time.strftime('%#m/%#d/%Y', time.localtime(td[0]))
            output = ','.join(td)
            f.write(output)
            f.write('\n')



'''GET THE OTHER OF THE DATA'''
files = [x.strip(r'.txt') for x in list(os.walk(path4))[0][-1]]
dropped_df = df_cramer[df_cramer['Ticker'].isin(files)].reset_index()
dropped_df.columns 
len(dropped_df)  # 11591
len(df_cramer)  # 12048



def analyze_calls():

    data_vector = list()

    idx = 0
    checker = set()
    ticker_data_struture = dict()    
    for t in dropped_df['Ticker']:
        
        ticker_vector = list() 
        print(idx, t)
        
        call_date = dropped_df['Date'].loc[idx]
        call_price = float(dropped_df['Price'].loc[idx])
        
        next_day_date = 0
        if t not in checker:
            print('{} opened file'.format(t))
            try:
                second_level_dict = dict()
                with open(path4+'\{}.txt'.format(t), 'r') as f:
                   read_in_data = [data.strip('\n').split(",") for data in f.readlines()]          
                    
                holder = [second_level_dict.update({x[0]:x[:]}) for x in read_in_data]
                keys_list = list(second_level_dict.keys())
                keys_list.sort(key = lambda date: datetime.datetime.strptime(date, '%m/%d/%Y')) 
                second_level_dict['all_keys'] = keys_list[::-1]
                ticker_data_struture[t] = [second_level_dict]
                checker.add(t)
            except Exception as e:
                print('{} --- {}'.format(t, e))   
        else:
            print('Did not open file')
        try:
            tick = ticker_data_struture[t][0]
            analyze_data = tick['all_keys']

            print('NEXT DATE ---', tick[analyze_data[analyze_data.index(call_date)-1]][0])
            
            # Dates
            next_day_date = tick[analyze_data[analyze_data.index(call_date)-1]][0] if analyze_data.index(call_date)-1 >= 0 else 'NAN'
            one_week_date = tick[analyze_data[analyze_data.index(call_date)-5]][0] if analyze_data.index(call_date)-5 >= 0 else 'NAN'
            one_month_date = tick[analyze_data[analyze_data.index(call_date)-20]][0] if analyze_data.index(call_date)-20 >= 0 else 'NAN'
            three_month_date = tick[analyze_data[analyze_data.index(call_date)-60]][0] if analyze_data.index(call_date)-60 >= 0 else 'NAN'
            
            # Prices
            next_day_closeprice = float(tick[analyze_data[analyze_data.index(call_date)-1]][4]) if next_day_date != 'NAN' else 'NAN'
            one_week_closeprice = float(tick[analyze_data[analyze_data.index(call_date)-5]][4]) if one_week_date != 'NAN' else 'NAN'
            one_month_closeprice = float(tick[analyze_data[analyze_data.index(call_date)-20]][4]) if one_month_date != 'NAN' else 'NAN'
            three_month_closeprice = float(tick[analyze_data[analyze_data.index(call_date)-60]][4]) if three_month_date != 'NAN' else 'NAN'
            
            # volume 
            next_day_volume = float(tick[analyze_data[analyze_data.index(call_date)-1]][5]) if next_day_date != 'NAN' else 'NAN'
            EMA10_volume = [int(tick[analyze_data[analyze_data.index(call_date)+int(v)]][5]) for v in range(1,11)]
            if len(EMA10_volume) != 10:
                raise ValueError('Volume less than 10 days')
                
                
            ema10_df = pd.DataFrame(EMA10_volume)
            volume_average = ema10_df.ewm(span=10,adjust=False).mean()[0][9]
            
            
            # percent chnage for price and volume 
            day_change = str(((next_day_closeprice/call_price)-1) * 100) if next_day_date != 'NAN' else 'NAN'
            week_change = str(((one_week_closeprice/call_price)-1) * 100) if one_week_date != 'NAN' else 'NAN'
            month_change = str(((one_month_closeprice/call_price)-1) * 100) if one_month_date != 'NAN' else 'NAN'
            threemonth_change = str(((three_month_closeprice/call_price)-1) * 100) if three_month_date != 'NAN' else 'NAN'
            volume_change = str(((next_day_volume/volume_average)-1) * 100) if next_day_date != 'NAN' else 'NAN'
            
            # measure of risk (annalized standard deviation)        
            small_df = [float(tick[x][4]) for x in analyze_data[analyze_data.index(call_date)+1:]][::-1]
            log_returns = np.diff(np.log(small_df)) if small_df != 'NAN' else 'NAN'
            sec_returns_std_annual = np.std(log_returns) * np.sqrt(250)
    
            # append data to data vector list 
            ticker_vector += ([next_day_date] + [next_day_closeprice] + [day_change] + 
                              [one_week_date] + [one_week_closeprice] + [week_change] + 
                              [one_month_date] + [one_month_closeprice] + [month_change] + 
                              [three_month_date] + [three_month_closeprice] + [threemonth_change] +
                              [volume_change] + [sec_returns_std_annual])                    
            

            data_vector.append(ticker_vector)
            
            if next_day_date == 0:
                raise ValueError('Date not found')
            else:
                idx += 1
                    
        except Exception as e:
            print(e)
            ticker_vector += ['NAN'] * 14
            idx += 1     
            data_vector.append(ticker_vector)
            continue

        
    data_df = pd.DataFrame(data_vector, columns=["Next_Day","Next_Day_Price","NextDay_Change",
                                                "1Week_Date","1Week_Price","1Week_Change",
                                                "1Month_Date","1Month_Price","1month_Change",
                                                "3Month_Date","3Month_Price","3month_Change",
                                                "NextDay_Volume_Change", "Annual_STD"])

    backtest_df = pd.concat([dropped_df,data_df], axis=1)
    backtest_df.to_csv('Backtest.csv')

analyze_calls()























